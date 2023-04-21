from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db import IntegrityError
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from .models import User, Choices, Questions, Answer, Form, Responses,District,Taluka,Block,Sector,AWC,Village
from django.core import serializers
import json
import random,re
import string
from datetime import datetime, date
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy

from django.core.paginator import Paginator
import openpyxl
import csv
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from .models import Form, Responses


def responsedetails(request, code, response_code):
    form = Form.objects.get(code=code)
    responses = Responses.objects.filter(response_code=response_code)

    answers_by_question = {}

    for response in responses:
        answers = Answer.objects.filter(response=response)
        for answer in answers:
            question_id = answer.answer_to.id
            question_text = answer.answer_to.question
            question_type = answer.answer_to.question_type
            print(question_text)
            if question_type == "districts":
                answer_text = District.objects.filter(id= answer.answer).first()
            elif question_type == "talukas":
                 answer_text = Taluka.objects.filter(id= answer.answer).first()
            elif question_type == "blocks":
                answer_text = Block.objects.filter(id= answer.answer).first()

            elif question_type == "sectors":
                answer_text = Sector.objects.filter(id= answer.answer).first()
            elif question_type == "villages":
                answer_text = Village.objects.filter(id= answer.answer).first()

            elif question_type == "awcs":
                answer_text = AWC.objects.filter(id= answer.answer).first()

            else: 
                answer_text = answer.answer
            if question_id in answers_by_question:
                answers_by_question[question_id]['answers'].append(answer_text)
            else:
                answers_by_question[question_id] = {'question_text': question_text,'question_type':question_type,'answers': [answer_text]}
            response = response
    context = {'form': form, 'answers_by_question': answers_by_question,'response_code':response_code ,'response':response}
    return render(request, 'index/edit_responsenew.html', context)


def view_data(request, code):

    form = get_object_or_404(Form, code=code)
    questions = form.questions.all().order_by('id')
    headers = [question.question for question in questions]
    data = []
    exportdata = []
    responseform = []
    for response in form.response_to.order_by('-id').all():
        row = []
        exportrow = []
       
        responseform.append(response.response_code)
        row.append(response.response_code)
        for question in questions:
            answer = response.response.filter(answer_to=question).first()
        
            if answer:
                
                if question.question_type == "districts" :
                    anscheck = answer.answer
                    # choice = answer.answer_to.choices.get(id = answer.answer).choice
                    if answer.answer !="":

                        choice = District.objects.filter(id=answer.answer).first()
                    else:
                        choice = District.objects.filter(id=0).first()
                    row.append(str(choice))
                    exportrow.append(str(choice))

                elif question.question_type == "talukas":
            
                    if answer.answer !="":

                        choice = Taluka.objects.filter(id=answer.answer).first()
                    else:
                        choice = Taluka.objects.filter(id=0).first()
                    row.append(str(choice))
                    exportrow.append(str(choice))
                    
                elif question.question_type == "blocks":
            
                    if answer.answer !="":

                        choice = Block.objects.filter(id=answer.answer).first()
                    else:
                        choice =""
                    row.append(str(choice))
                    exportrow.append(str(choice))
                    

                elif question.question_type == "sectors":
            
                    if answer.answer !="":

                        choice = Sector.objects.filter(id=answer.answer).first()
                    else:
                        choice =""
                    row.append(str(choice))
                    exportrow.append(str(choice))
                    
                elif question.question_type == "villages":
            
                    if answer.answer !="":

                        choice = Village.objects.filter(id=answer.answer).first()
                    else:
                        choice =""
                    row.append(str(choice))
                    exportrow.append(str(choice))
                    

                elif question.question_type == "awcs":
            
                    if answer.answer !="":

                        choice = AWC.objects.filter(id=answer.answer).first()
                    else:
                        choice =""
                    row.append(str(choice))
                    exportrow.append(str(choice))

                elif question.question_type == "date":
            
                    choice = answer.answer

                    print("date from databse")
                    print(choice)
                    choice  = re.sub(r'(\d{4})-(\d{1,2})-(\d{1,2})', '\\3-\\2-\\1', choice)
                    # formatDate = datetime(choice)
                    # formatDate = formatDate.strftime("%d/%B/%Y")
                    # print("new date formqte from databse")
                    # print(formatDate)
                    row.append(str(choice))
                    exportrow.append(str(choice))
                    
                    

                elif question.question_type == "multiple choice" or question.question_type == "checkbox":
            
                    # choice = answer.answer_to.choices.get(id = answer.answer).choice
                    
                    choice = answer.answer_to.choices.get(id = answer.answer).choice
                   
                    row.append(choice)
                    exportrow.append(choice)

                else:
                    
                    row.append(answer.answer)
                    exportrow.append(answer.answer)
                    
                    
            else:
                row.append('')
                exportrow.append('')
        data.append(row)

        exportdata.append(exportrow)
        print("printing data")
        print(data)
    if 'export' in request.POST:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="responses.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Responses'
        for index, header in enumerate(headers):
            print(index)
            worksheet.cell(row=1, column=index+1, value=header)

            
        for row_index, exportrow in enumerate(exportdata):
            for column_index, cell in enumerate(exportrow):
                print("printing cell")
                print(cell)
                worksheet.cell(row=row_index+2, column=column_index+1, value=cell)
        workbook.save(response)
        return response

    return render(request, 'index/form_newresponse.html', {'form': form, 'headers': headers,'response':responseform, 'data': data})
    # return render(request, 'index/form_newresponse.html', context)


    # return render(request, 'index/form_newresponse.html', context)
# Create your views here.
def index(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    forms = Form.objects.filter(creator = request.user)
    return render(request, "index/index.html", {
        "forms": forms
    })
def home(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('home'))
    forms = Form.objects.filter(creator = request.user)
    return render(request, "index/index.html", {
        "forms": forms
    })
# def blockes(request,code):
#     blockes = Block.objects.all()
#     context = {'blockes' : blockes,'form': formInfo,'code': code,
#           }
    
#     # district_id = request.GET.get('district')
#     # blockes = Block.objects.filter(district_id=district_id).order_by('name')
#     # return render(request, 'hr/city_dropdown_list_options.html', {'blockes': blockes})
#     formInfo = Form.objects.filter(code = code)
#     return render(request, "index/block_dropdown_list_options.html", {
#             "code": code,
#             "form": formInfo,
#             "blockes": blockes

#         })



def login_view(request):
    #Check if the user is logged in
    print("printing users nature : " + str(request.user.is_authenticated))
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('index'))
    if request.method == "POST":
        username = request.POST["username"].lower()
        password = request.POST["password"]
        user = authenticate(request, username = username, password = password)
        # if user authentication success
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse('index'))
        else:
            return render(request, "index/login.html", {
                "message": "Invalid username and/or password"
            })
    return render(request, "index/login.html")

def howtocreate(request):
    return render(request, "index/howtocreate.html")


def register(request):
    #Check if the user is logged in
    if request.user.is_authenticated:
        
        return HttpResponseRedirect(reverse('index'))
    if request.method == "POST":
        username = request.POST["username"].lower()
        password = request.POST["password"]
        email = request.POST["email"]
        confirmation = request.POST["confirmation"]
        #check if the password is the same as confirmation
        if password != confirmation:
            return render(request, "index/register.html", {
                "message": "Passwords must match."
            })
        #Checks if the username is already in use
        if User.objects.filter(email = email).count() == 1:
            return render(request, "index/register.html", {
                "message": "Email already taken."
            })
        try:
            user = User.objects.create_user(username = username, password = password, email = email)
            user.save()
            login(request, user)
            return HttpResponseRedirect(reverse('index'))
        except IntegrityError:
            return render(request, "index/register.html", {
                "message": "Username already taken"
            })
    return render(request, "index/register.html")


def logout_view(request):
    #Logout the user
    logout(request)
    return HttpResponseRedirect(reverse('index'))

def create_form(request):
    # Creator must be authenticated
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    # Create a blank form API
    if request.method == "POST":
        data = json.loads(request.body)
        title = data["title"]
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        choices = Choices(choice = "Option 1")
        choices.save()
        question = Questions(question_type = "multiple choice", question= "Untitled Question", required= False)
        question.save()
        question.choices.add(choices)
        question.save()
        form = Form(code = code, title = title, creator=request.user)
        form.save()
        form.questions.add(question)
        form.save()
        return JsonResponse({"message": "Sucess", "code": code})

def district_form(request):
    # Creator must be authenticated
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    # Create a blank form API
    if request.method == "POST":
        data = json.loads(request.body)
        title = data["title"]
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        districtlist = District.objects.all()
        # choices.save()
        question = Questions(question_type = "districs", question= "District", required= False)
        question.save()
        # question.choices.add(choices)
        question.save()
        form = Form(code = code,is_district = 1, title = title, creator=request.user)
        form.save()
        form.questions.add(question)
        form.save()
        return JsonResponse({"message": "Sucess", "code": code, "districtlist":districtlist})




def edit_form(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    
    if formInfo.is_district:
        print("this district level form")
        datalist = District.objects.all()
    else:
        datalist={}    
    print("edit form method call")
    print(request)

    
    return render(request, "index/form.html", {
        "code": code,
        "form": formInfo,
        "datalist" :datalist
    })
def load_cities(request,code):
    district_id = request.GET.get('district')
    blockes = Block.objects.filter(district_id=district_id).order_by('name')
    # return render(request, 'hr/city_dropdown_list_options.html', {'blockes': blockes})
    formInfo = Form.objects.filter(code = code)
    return render(request, "index/block_dropdown_list_options.html", {
            "code": code,
            "form": formInfo,
            "blockes": blockes

        })
def edit_title(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        if len(data["title"]) > 0:
            formInfo.title = data["title"]
            formInfo.save()
        else:
            formInfo.title = formInfo.title[0]
            formInfo.save()
        return JsonResponse({"message": "Success", "title": formInfo.title})

def edit_description(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.description = data["description"]
        formInfo.save()
        return JsonResponse({"message": "Success", "description": formInfo.description})

def edit_bg_color(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.background_color = data["bgColor"]
        formInfo.save()
        return JsonResponse({"message": "Success", "bgColor": formInfo.background_color})

def edit_text_color(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.text_color = data["textColor"]
        formInfo.save()
        return JsonResponse({"message": "Success", "textColor": formInfo.text_color})

def edit_setting(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        formInfo.collect_email = data["collect_email"]
        formInfo.is_quiz = data["is_quiz"]
        formInfo.authenticated_responder = data["authenticated_responder"]
        formInfo.confirmation_message = data["confirmation_message"]
        formInfo.edit_after_submit = data["edit_after_submit"]
        formInfo.allow_view_score = data["allow_view_score"]
        formInfo.save()
        return JsonResponse({'message': "Success"})

def delete_form(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse("404"))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":
        #Delete all questions and choices
        for i in formInfo.questions.all():
            for j in i.choices.all():
                j.delete()
            i.delete()
        for i in Responses.objects.filter(response_to = formInfo):
            for j in i.response.all():
                j.delete()
            i.delete()
        formInfo.delete()
        return JsonResponse({'message': "Success"})

def edit_question(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        question_id = data["id"]
        question = Questions.objects.filter(id = question_id)
        if question.count() == 0:
            return HttpResponseRedirect(reverse("404"))
        else: question = question[0]
        question.question = data["question"]
        question.question_type = data["question_type"]
        question.required = data["required"]
        if(data.get("score")): question.score = data["score"]
        if(data.get("answer_key")): question.answer_key = data["answer_key"]
        question.save()
        return JsonResponse({'message': "Success"})

def edit_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        choice_id = data["id"]
        choice = Choices.objects.filter(id = choice_id)
        if choice.count() == 0:
            return HttpResponseRedirect(reverse("404"))
        else: choice = choice[0]
        choice.choice = data["choice"]
        if(data.get('is_answer')): choice.is_answer = data["is_answer"]
        choice.save()
        return JsonResponse({'message': "Success"})

def add_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    
    if request.method == "POST":
        data = json.loads(request.body)
        choice = Choices(choice="Option")
        print("Printing Data form add choise methosd")
        print(data)
        choice.save()
        formInfo.questions.get(pk = data["question"]).choices.add(choice)
        formInfo.save()
        return JsonResponse({"message": "Success", "choice": choice.choice, "id": choice.id})

def add_district(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        district = District(name="Option")
        print("Printing Data form add choise methosd")
        print(district)
        district.save()
        formInfo.questions.get(pk = data["question"]).districts.add(district)
        print(formInfo.save())
        return JsonResponse({"message": "Success", "choice": district.name, "id": district.id})


def remove_choice(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        data = json.loads(request.body)
        choice = Choices.objects.filter(pk = data["id"])
        if choice.count() == 0:
            return HttpResponseRedirect(reverse("404"))
        else: choice = choice[0]
        choice.delete()
        return JsonResponse({"message": "Success"})

def get_choice(request, code, question):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "GET":
        question = Questions.objects.filter(id = question)
        if question.count() == 0: return HttpResponseRedirect(reverse('404'))
        else: question = question[0]
        choices = question.choices.all()
        choices = [{"choice":i.choice, "is_answer":i.is_answer, "id": i.id} for i in choices]
        return JsonResponse({"choices": choices, "question": question.question, "question_type": question.question_type, "question_id": question.id})

def add_question(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "POST":
        choices = Choices(choice = "Option 1")
        choices.save()
        question = Questions(question_type = "multiple choice", question= "Untitled Question", required= False)
        question.save()
        question.choices.add(choices)
        question.save()
        formInfo.questions.add(question)
        formInfo.save()
        return JsonResponse({'question': {'question': "Untitled Question", "question_type": "multiple choice", "required": False, "id": question.id}, 
        "choices": {"choice": "Option 1", "is_answer": False, 'id': choices.id}})

def delete_question(request, code, question):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":
        question = Questions.objects.filter(id = question)
        if question.count() == 0: return HttpResponseRedirect(reverse("404"))
        else: question = question[0]
        for i in question.choices.all():
            i.delete()
            question.delete()
        return JsonResponse({"message": "Success"})

def score(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        return render(request, "index/score.html", {
            "form": formInfo
        })

def edit_score(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question_id = data["question_id"]
            question = formInfo.questions.filter(id = question_id)
            if question.count() == 0:
                return HttpResponseRedirect(reverse("edit_form", args = [code]))
            else: question = question[0]
            score = data["score"]
            if score == "": score = 0
            question.score = score
            question.save()
            return JsonResponse({"message": "Success"})

def answer_key(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question = Questions.objects.filter(id = data["question_id"])
            if question.count() == 0: return HttpResponseRedirect(reverse("edit_form", args = [code]))
            else: question = question[0]
            if question.question_type == "short" or question.question_type == "paragraph":
                question.answer_key = data["answer_key"]
                question.save()
            else:
                for i in question.choices.all():
                    i.is_answer = False
                    i.save()
                if question.question_type == "multiple choice":
                    choice = question.choices.get(pk = data["answer_key"])
                    choice.is_answer = True
                    choice.save()
                else:
                    for i in data["answer_key"]:
                        choice = question.choices.get(id = i)
                        choice.is_answer = True
                        choice.save()
                question.save()
            return JsonResponse({'message': "Success"})

def feedback(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if not formInfo.is_quiz:
        return HttpResponseRedirect(reverse("edit_form", args = [code]))
    else:
        if request.method == "POST":
            data = json.loads(request.body)
            question = formInfo.questions.get(id = data["question_id"])
            question.feedback = data["feedback"]
            question.save()
            return JsonResponse({'message': "Success"})

def view_form(request, code):

    formInfo = Form.objects.filter(code = code)
    
    
    districtname= District.objects.all()
    blockname= {}
    talukaname = Taluka.objects.all()
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if formInfo.authenticated_responder:
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse("login"))
    
    context = {'form': formInfo,'code': code, 'districts' : districtname, 'talukas' : talukaname
          }
    return render(request, "index/view_form.html", context)

def talukas(request):
    print("taluka methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        
    district= request.GET.get('districts')

    
    taluka = Taluka.objects.filter(district= data[0])
    
    context = {"taluka" : taluka
          }
    return render(request, "partials/taluka.html", context) 

def edittalukas(request):
    print("taluka methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    district= request.GET.get('districts')

    
    talukas = Taluka.objects.filter(district = data[0])
    print(talukas)
    context = {"talukas" : talukas
          }
    return render(request, "partials/edittaluka.html", context) 





def blockes(request):
    print("Block methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        
        taluka= request.GET.get('talukas')

        
        blockes = Block.objects.filter(taluka= data[0])
        
    context = {"blockes" : blockes
          }
    return render(request, "partials/block.html", context) 

def editblockes(request):
    print("editBlock methos call")
    if request.htmx:
        print("rewust is htmx")
        print(request.GET)
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        
        taluka= request.GET.get('talukas')

        
        blockes = Block.objects.filter(taluka= data[0])
        
    context = {"blockes" : blockes
          }
    return render(request, "partials/editblock.html", context) 

def sectors(request):
    print("Sector methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    block= request.GET.get('blocks')

    
    sectors = Sector.objects.filter(block = data[0])
    
    context = {"sectors" : sectors
          }
    return render(request, "partials/sector.html", context) 

def editsectors(request):
    print("Sector methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    block= request.GET.get('blocks')

    
    sectors = Sector.objects.filter(block = data[0])
    
    context = {"sectors" : sectors
          }
    return render(request, "partials/editsector.html", context) 

def editvillages(request):
    print("village methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    sectors= request.GET.get('sectors')

    
    villages = Village.objects.filter(block = data[0])
    
    context = {"villages" : villages
          }
    return render(request, "partials/editvillage.html", context) 


def awcs(request):
    print("awc methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    sector= request.GET.get('sectors')

    print(sector)
    awcs = AWC.objects.filter(sector= data[0])
    print(awcs)
    context = {"awcs" : awcs
          }
    return render(request, "partials/awc.html", context) 
def editawcs(request):
    print("awc methos call")
    if request.htmx:
        
        for i in request.GET:
            print("check values in GET methos")
            print(i)

        data = request.GET.getlist(i)
        print("printing the id which is unkone name")
        print(data)
    sector= request.GET.get('sectors')

    print(sector)
    awcs = AWC.objects.filter(sector= data[0])
    print(awcs)
    context = {"awcs" : awcs
          }
    return render(request, "partials/editawc.html", context) 

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip





def submit_form(request, code):
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if formInfo.authenticated_responder:
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse("login"))
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(20))
        if formInfo.authenticated_responder:
            response = Responses(response_code = code, response_to = formInfo, responder_ip = get_client_ip(request), responder = request.user)
            response.save()
        else:
            if not formInfo.collect_email:
                response = Responses(response_code = code, response_to = formInfo, responder_ip = get_client_ip(request))
                response.save()
            else:
                response = Responses(response_code = code, response_to = formInfo, responder_ip = get_client_ip(request), responder_email=request.POST["email-address"])
                response.save()
        for i in request.POST:
            #Excluding csrf token
            if i == "csrfmiddlewaretoken" or i == "email-address":
                continue
            if i=="districts":
                question = formInfo.questions.get(question_type = i)
                
            elif i=="talukas":
                question = formInfo.questions.get(question_type = i)
            elif i=="blocks":
                question = formInfo.questions.get(question_type = i)
            elif i=="sectors":
                question = formInfo.questions.get(question_type = i)
            elif i=="awcs":
                question = formInfo.questions.get(question_type = i)
                
            else:

                question = formInfo.questions.get(id = i)
               
            for j in request.POST.getlist(i):
                
                answer = Answer(answer=j, answer_to = question)
                answer.save()
                response.response.add(answer)
                response.save()
            
               
        return render(request, "index/form_response.html", {
            "form": formInfo,
            "code": code,
           
        })

def responses(request, code):
    print("Response method called")
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]

    responsesSummary = []
    choiceAnswered = {}
    districtAnswered= {}
    talukaAnswered = {}
    blockAnswered = {}
    sectorAnswered = {}
    villageAnswered = {}
    awcAnswered = {}
    filteredResponsesSummary = {}
    answerResponsesSummary = {}
    
    for question in formInfo.questions.all():

        answers = Answer.objects.filter(answer_to = question.id)
       
        
        if question.question_type == "multiple choice" or question.question_type == "checkbox" :
            choiceAnswered[question.question] = choiceAnswered.get(question.question, {})
            
            
            for answer in answers:
                
                choice = answer.answer_to.choices.get(id = answer.answer).choice
                

                choiceAnswered[question.question][choice] = choiceAnswered.get(question.question, {}).get(choice, 0) + 1

            
            
        
        elif question.question_type == "districts" :
            print("hello from district ansers")
            districtAnswered[question.question] = districtAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = District.objects.filter(id=answer.answer).first()
                
                districtAnswered[question.question][choice] = districtAnswered.get(question.question, {}).get(choice, 0) + 1
            
        elif question.question_type == "talukas" :
            
            talukaAnswered[question.question] = talukaAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = Taluka.objects.filter(id=answer.answer).first()
                talukaAnswered[question.question][choice] = talukaAnswered.get(question.question, {}).get(choice, 0) + 1
            
        elif question.question_type == "blocks" :
            
            blockAnswered[question.question] = blockAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = Block.objects.filter(id=answer.answer).first() 
                blockAnswered[question.question][choice] = blockAnswered.get(question.question, {}).get(choice, 0) + 1
            
        elif question.question_type == "sectors" :
            
            sectorAnswered[question.question] = sectorAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = Sector.objects.filter(id=answer.answer).first()
                sectorAnswered[question.question][choice] = sectorAnswered.get(question.question, {}).get(choice, 0) + 1
            
        elif question.question_type == "villages" :
            print("hello from district ansers")
            villageAnswered[question.question] = villageAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = Village.objects.filter(id=answer.answer).first()
                villageAnswered[question.question][choice] = villageAnswered.get(question.question, {}).get(choice, 0) + 1
            
        elif question.question_type == "awcs" :
            print("hello from awcAnswered ansers")
            awcAnswered[question.question] = awcAnswered.get(question.question, {})
            
            for answer in answers:
                if answer.answer=="" :

                    choice = ""
                else:
                

                    choice = AWC.objects.filter(id=answer.answer).first()
                awcAnswered[question.question][choice] = awcAnswered.get(question.question, {}).get(choice, 0) + 1
            
        responsesSummary.append({"question": question, "answers":answers })
    for answr in choiceAnswered:
        filteredResponsesSummary[answr] = {}
        keys = choiceAnswered[answr].values()
        for choice in choiceAnswered[answr]:
            filteredResponsesSummary[answr][choice] = choiceAnswered[answr][choice]
            
   
    for answr in districtAnswered:
        answerResponsesSummary[answr] = {}
        keys = districtAnswered[answr].values()
        for choice in districtAnswered[answr]:
            answerResponsesSummary[answr][choice] = districtAnswered[answr][choice]
            
    for answr in talukaAnswered:
        answerResponsesSummary[answr] = {}
        keys = talukaAnswered[answr].values()
        for choice in talukaAnswered[answr]:
            answerResponsesSummary[answr][choice] = talukaAnswered[answr][choice]
            
    for answr in blockAnswered:
        answerResponsesSummary[answr] = {}
        keys = blockAnswered[answr].values()
        for choice in blockAnswered[answr]:
            answerResponsesSummary[answr][choice] = blockAnswered[answr][choice]
    for answr in sectorAnswered:
        answerResponsesSummary[answr] = {}
        keys = sectorAnswered[answr].values()
        for choice in sectorAnswered[answr]:
            answerResponsesSummary[answr][choice] = sectorAnswered[answr][choice]
    for answr in awcAnswered:
        answerResponsesSummary[answr] = {}
        keys = awcAnswered[answr].values()
        for choice in awcAnswered[answr]:
            answerResponsesSummary[answr][choice] = awcAnswered[answr][choice]
    
        
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    
    print("rendering page")
    return render(request, "index/responses.html", {
        "form": formInfo,
        "responses": Responses.objects.filter(response_to = formInfo),
        "responsesSummary": responsesSummary,
        "filteredResponsesSummary": filteredResponsesSummary,
        "answerResponsesSummary" : answerResponsesSummary
        
    })

def response(request, code, response_code):
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if not formInfo.allow_view_score:
        if formInfo.creator != request.user:
            return HttpResponseRedirect(reverse("403"))
    total_score = 0
    score = 0
    responseInfo = Responses.objects.filter(response_code = response_code)
    if responseInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: responseInfo = responseInfo[0]
    if formInfo.is_quiz:
        for i in formInfo.questions.all():
            total_score += i.score
        for i in responseInfo.response.all():
            if i.answer_to.question_type == "short" or i.answer_to.question_type == "paragraph":
                if i.answer == i.answer_to.answer_key: score += i.answer_to.score
            elif i.answer_to.question_type == "multiple choice":
                answerKey = None
                for j in i.answer_to.choices.all():
                    if j.is_answer: answerKey = j.id
                if answerKey is not None and int(answerKey) == int(i.answer):
                    score += i.answer_to.score
        _temp = []
        for i in responseInfo.response.all():
            if i.answer_to.question_type == "checkbox" and i.answer_to.pk not in _temp:
                answers = []
                answer_keys = []
                for j in responseInfo.response.filter(answer_to__pk = i.answer_to.pk):
                    answers.append(int(j.answer))
                    for k in j.answer_to.choices.all():
                        if k.is_answer and k.pk not in answer_keys: answer_keys.append(k.pk)
                    _temp.append(i.answer_to.pk)
                if answers == answer_keys: score += i.answer_to.score
    return render(request, "index/response.html", {
        "form": formInfo,
        "response": responseInfo,
        "score": score,
        "total_score": total_score
    })

def edit_response(request, code, response_code):
    formInfo = Form.objects.filter(code = code)
    
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    response = Responses.objects.filter(response_code = response_code, response_to = formInfo)
    
    # choice = District.objects.filter(id=answer.answer).first()
    if response.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: response = response[0]
    if formInfo.authenticated_responder:
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse("login"))
        if response.responder != request.user:
            return HttpResponseRedirect(reverse('403'))

    # Accessing answer for the question type "districts"
    alldistrict = District.objects.all()
    district_answer = Answer.objects.filter(answer_to__question_type='districts', response=response).first()
    district = None

    # Checking if answer exists for the question type "districts"
    if district_answer:
        district_id = district_answer.answer
        print("district id is blacnk")
        print(district_answer)
        if district_id=="":
            print("district id is blacnk")
            print(district_answer)
            district= ""
        else: 
            district = District.objects.filter(id=district_id).first()

     # Accessing answer for the question type "taluka"
    
    taluka_answer = Answer.objects.filter(answer_to__question_type='talukas', response=response).first()
    taluka = None
    print("taluka name none : ")
    print(taluka)
    # Checking if answer exists for the question type "taluka"
    if taluka_answer:
        taluka_id = taluka_answer.answer
        if taluka_id=="":
            
      
            taluka= ""
        else:
            taluka = Taluka.objects.filter(id=taluka_id).first()
            

    # Accessing answer for the question type "block"
    allblock = Block.objects.all()
    block_answer = Answer.objects.filter(answer_to__question_type='blocks', response=response).first()
    block = None
    
    # Checking if answer exists for the question type "block"
    if block_answer:
        block_id = block_answer.answer
        if block_id=="":
            block= ""
        else:
            block = Block.objects.filter(id=block_id).first()
        
    print(block)

     # Accessing answer for the question type "sector"
    allsector = Sector.objects.all()
    sector_answer = Answer.objects.filter(answer_to__question_type='sectors', response=response).first()
    sector = None
   


     # Checking if answer exists for the question type "sector"
    if sector_answer:
        sector_id = sector_answer.answer
        if sector_id=="":
            
            sector= ""
        else:
            sector = Block.objects.filter(id=sector_id).first()
        
     # Accessing answer for the question type "village"
    allvillage = Village.objects.all()
    village_answer = Answer.objects.filter(answer_to__question_type='villages', response=response).first()
    village = None
     # Checking if answer exists for the question type "village"
    if village_answer:
        village_id = village_answer.answer
        if village_id=="":
            village= ""
        else:
            village = Village.objects.filter(id=village_id).first()
        print("block name")
        print(village)
    print(village)

     # Accessing answer for the question type "awcs"
    allawc = AWC.objects.all()
    awc_answer = Answer.objects.filter(answer_to__question_type='awcs', response=response).first()
    awc = None
    print("village name none : ")
    print(awc)


     # Checking if answer exists for the question type "awcs"
    if awc_answer:
        awc_id = awc_answer.answer
        if awc_id=="":
            print("id is blacnk")
            print(awc_id)
      
            awc= ""
        else:
            awc = AWC.objects.filter(id=awc_id).first()
        print("block name")
        print(awc)
    print(awc)

    # below is if response edited and sumbimted for data save 
    if request.method == "POST":
        if formInfo.authenticated_responder and not response.responder:
            response.responder = request.user
            response.save()
        if formInfo.collect_email:
            response.responder_email = request.POST["email-address"]
            response.save()
        #Deleting all existing answers
        for i in response.response.all():
            i.delete()
        for i in request.POST:
            #Excluding csrf token and email address
            if i == "csrfmiddlewaretoken" or i == "email-address":
                continue
            # check type for district taluka block sector
            if i=="districts":
                question = formInfo.questions.get(question_type = i)
                
            elif i=="talukas":
                question = formInfo.questions.get(question_type = i)
            elif i=="blocks":
                question = formInfo.questions.get(question_type = i)
            elif i=="sectors":
                question = formInfo.questions.get(question_type = i)
            elif i=="villages":
                question = formInfo.questions.get(question_type = i)
            elif i=="awcs":
                question = formInfo.questions.get(question_type = i)
                
            else:

                question = formInfo.questions.get(id = i)

            # question = formInfo.questions.get(id = i)
            for j in request.POST.getlist(i):
                answer = Answer(answer=j, answer_to = question)
                answer.save()
                response.response.add(answer)
                response.save()
        if formInfo.is_quiz:
            return HttpResponseRedirect(reverse("response", args = [formInfo.code, response.response_code]))
        else:
            return render(request, "index/form_response.html", {
                "form": formInfo,
                "code": response.response_code
            })
    return render(request, "index/edit_responsenew.html", {
        "form": formInfo,
        "response": response,
        "district":district,
        "alldistrict" :alldistrict,
        "taluka" :taluka,
        "blocked" : block,
        "sector" : sector,
        "village" : village,
        "awc" : awc

    })

def contact_form_template(request):
    # Creator must be authenticated
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    # Create a blank form API
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        name = Questions(question_type = "short", question= "Name", required= True)
        name.save()
        email = Questions(question_type="short", question = "Email", required = True)
        email.save()
        address = Questions(question_type="paragraph", question="Address", required = True)
        address.save()
        phone = Questions(question_type="short", question="Phone number", required = False)
        phone.save()
        comments = Questions(question_type = "paragraph", question = "Comments", required = False)
        comments.save()
        form = Form(code = code, title = "Contact information", creator=request.user, background_color="#e2eee0", allow_view_score = False, edit_after_submit = True)
        form.save()
        form.questions.add(name)
        form.questions.add(email)
        form.questions.add(address)
        form.questions.add(phone)
        form.questions.add(comments)
        form.save()
        return JsonResponse({"message": "Sucess", "code": code})




def customer_feedback_template(request):
    # Creator must be authenticated
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    # Create a blank form API
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        comment = Choices(choice = "Comments")
        comment.save()
        question = Choices(choice = "Questions")
        question.save()
        bug = Choices(choice = "Bug Reports")
        bug.save()
        feature = Choices(choice = "Feature Request")
        feature.save()
        feedback_type = Questions(question = "Feedback Type", question_type="multiple choice", required=False)
        feedback_type.save()
        feedback_type.choices.add(comment)
        feedback_type.choices.add(bug)
        feedback_type.choices.add(question)
        feedback_type.choices.add(feature)
        feedback_type.save()
        feedback = Questions(question = "Feedback", question_type="paragraph", required=True)
        feedback.save()
        suggestion = Questions(question = "Suggestions for improvement", question_type="paragraph", required=False)
        suggestion.save()
        name = Questions(question = "Name", question_type="short", required=False)
        name.save()
        email = Questions(question= "Email", question_type="short", required=False)
        email.save()
        form = Form(code = code, title = "Customer Feedback", creator=request.user, background_color="#e2eee0", confirmation_message="Thanks so much for giving us feedback!",
        description = "We would love to hear your thoughts or feedback on how we can improve your experience!", allow_view_score = False, edit_after_submit = True)
        form.save()
        form.questions.add(feedback_type)
        form.questions.add(feedback)
        form.questions.add(suggestion)
        form.questions.add(name)
        form.questions.add(email)
        return JsonResponse({"message": "Sucess", "code": code})

def event_registration_template(request):
    # Creator must be authenticated
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    # Create a blank form API
    if request.method == "POST":
        code = ''.join(random.choice(string.ascii_letters + string.digits) for x in range(30))
        name = Questions(question="Name", question_type= "short", required=False)
        name.save()
        email = Questions(question = "email", question_type="short", required=True)
        email.save()
        organization = Questions(question = "Organization", question_type= "short", required=True)
        organization.save()
        day1 = Choices(choice="Day 1")
        day1.save()
        day2 = Choices(choice= "Day 2")
        day2.save()
        day3 = Choices(choice= "Day 3")
        day3.save()
        day = Questions(question="What days will you attend?", question_type="checkbox", required=True)
        day.save()
        day.choices.add(day1)
        day.choices.add(day2)
        day.choices.add(day3)
        day.save()
        dietary_none = Choices(choice="None")
        dietary_none.save()
        dietary_vegetarian = Choices(choice="Vegetarian")
        dietary_vegetarian.save()
        dietary_kosher = Choices(choice="Kosher")
        dietary_kosher.save()
        dietary_gluten = Choices(choice = "Gluten-free")
        dietary_gluten.save()
        dietary = Questions(question = "Dietary restrictions", question_type="multiple choice", required = True)
        dietary.save()
        dietary.choices.add(dietary_none)
        dietary.choices.add(dietary_vegetarian)
        dietary.choices.add(dietary_gluten)
        dietary.choices.add(dietary_kosher)
        dietary.save()
        accept_agreement = Choices(choice = "Yes")
        accept_agreement.save()
        agreement = Questions(question = "I understand that I will have to pay $$ upon arrival", question_type="checkbox", required=True)
        agreement.save()
        agreement.choices.add(accept_agreement)
        agreement.save()
        form = Form(code = code, title = "Event Registration", creator=request.user, background_color="#fdefc3", 
        confirmation_message="We have received your registration.\n\
Insert other information here.\n\
\n\
Save the link below, which can be used to edit your registration up until the registration closing date.",
        description = "Event Timing: January 4th-6th, 2016\n\
Event Address: 123 Your Street Your City, ST 12345\n\
Contact us at (123) 456-7890 or no_reply@example.com", edit_after_submit=True, allow_view_score=False)
        form.save()
        form.questions.add(name)
        form.questions.add(email)
        form.questions.add(organization)
        form.questions.add(day)
        form.questions.add(dietary)
        form.questions.add(agreement)
        form.save()
        return JsonResponse({"message": "Sucess", "code": code})

def delete_responses(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    #Checking if form exists
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    #Checking if form creator is user
    if formInfo.creator != request.user:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":
        responses = Responses.objects.filter(response_to = formInfo)
        for response in responses:
            for i in response.response.all():
                i.delete()
            response.delete()
        return JsonResponse({"message": "Success"})

# Error handler
def FourZeroThree(request):
    return render(request, "error/403.html")

def FourZeroFour(request):
    return render(request, "error/404.html")
